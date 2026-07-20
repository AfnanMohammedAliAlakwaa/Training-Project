from decimal import Decimal
import os
from urllib.parse import urlencode

from django.db import DatabaseError
from django.db.models import Max
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

import arabic_reshaper
from bidi.algorithm import get_display

from .models import (
    AcademicProgram,
    EvaluationFile,
    QualityStandard,
    StandardEntry,
    EvidenceAttachment,
    ReportExportLog,
)
from evaluations.models import ProgramEvaluationReview, StandardEvaluationReview
from improvements.models import ImprovementPlan


DEFAULT_COLLEGE_NAME = "غير محدد"
DEFAULT_ACADEMIC_YEAR = "all"
ALL_VALUES = {"", "all", "اختار البرنامج", "جميع البرامج", "جميع الكليات", "جميع الأنواع", "جميع السنوات"}

REPORT_DEFINITIONS = [
    {
        "id": "comprehensive",
        "title": "التقرير الشامل للبرنامج",
        "description": "تقرير شامل يعرض ملخص التقويم الذاتي للبرنامج الأكاديمي.",
    },
    {
        "id": "evaluation",
        "title": "تقرير نتائج التقييم",
        "description": "يعرض نتائج تقييم معايير البرنامج الأكاديمي.",
    },
    {
        "id": "attachments",
        "title": "تقرير المرفقات والشواهد",
        "description": "يعرض حالة الملاحق والشواهد المطلوبة للبرنامج.",
    },
    {
        "id": "strengths_weaknesses",
        "title": "تقرير نقاط القوة والضعف",
        "description": "يوضح نقاط القوة والضعف حسب نتائج التقييم.",
    },
    {
        "id": "improvement",
        "title": "تقرير خطط التحسين",
        "description": "يعرض خطط التحسين المفتوحة والمنجزة.",
    },
    {
        "id": "executive",
        "title": "التقرير التنفيذي",
        "description": "ملخص تنفيذي مختصر لأهم نتائج التقارير.",
    },
]

REPORT_TITLES = {item["id"]: item["title"] for item in REPORT_DEFINITIONS}

ACADEMIC_YEAR_START = 2007
ACADEMIC_YEAR_END = 2027

# هذه ليست بيانات تقارير وهمية؛ هذه قاعدة عمل ثابتة لبداية كل برنامج أكاديمي.
# تستخدم فقط لضبط قائمة السنوات الأكاديمية حسب البرنامج المختار.
PROGRAM_START_YEARS = {
    "هندسه البرمجيات": 2007,
    "هندسه برمجيات": 2007,
    "نظم معلومات - اعمال الكترونيه": 2008,
    "نظم معلومات اعمال الكترونيه": 2008,
    "اعمال الكترونيه": 2008,
    "تقنيه معلومات": 2011,
    "تقنيه معلومات انجليزي": 2019,
    "امن سيبراني": 2021,
    "نظم معلومات - ذكاء اعمال": 2021,
    "نظم معلومات ذكاء اعمال": 2021,
    "ذكاء اعمال": 2021,
    "ذكاء اصطناعي": 2024,
}


def _safe_text(value, default=""):
    if value is None:
        return default

    text = str(value).strip()
    if text.lower() in {"none", "null", "nan"}:
        return default

    return text


def _normalize_arabic(value):
    text = _safe_text(value)
    replacements = {
        "أ": "ا",
        "إ": "ا",
        "آ": "ا",
        "ٱ": "ا",
        "ة": "ه",
        "ى": "ي",
        "ؤ": "و",
        "ئ": "ي",
        "ـ": "",
        "–": "-",
        "—": "-",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    return " ".join(text.lower().split())


def _is_all_value(value):
    return _safe_text(value) in ALL_VALUES


def _to_float(value, default=0.0):
    try:
        if value is None:
            return default
        if isinstance(value, Decimal):
            return float(value)
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_int_percent(value):
    value = _to_float(value, 0)
    value = max(0, min(100, value))
    return int(round(value))


def _avg(values, default=0):
    valid_values = [_to_float(value) for value in values if value is not None]
    if not valid_values:
        return default
    return sum(valid_values) / len(valid_values)


def _format_score(value):
    number = _to_float(value, 0)
    if number <= 0:
        return "-"
    return round(number, 2)


def _format_date(value):
    if not value:
        return "لا توجد بيانات"

    try:
        return timezone.localtime(value).strftime("%Y/%m/%d")
    except Exception:
        try:
            return value.strftime("%Y/%m/%d")
        except Exception:
            return _safe_text(value, "لا توجد بيانات")


def _display_year(value):
    return _safe_text(value).replace("-", "/")


def _status_from_progress(progress):
    progress = _to_int_percent(progress)

    if progress >= 85:
        return "جاهز", "success"
    if progress >= 60:
        return "قيد المراجعة", "info"
    if progress > 0:
        return "مسودة", "warning"
    return "لا توجد بيانات", "danger"


def _standard_status(completion, score=None):
    completion = _to_int_percent(completion)
    score_value = _to_float(score, 0)

    if completion >= 90 and score_value >= 4:
        return "مكتمل"
    if completion >= 75:
        return "مكتمل جزئيًا"
    if completion >= 50:
        return "بحاجة مراجعة"
    if completion > 0:
        return "بحاجة تحسين"
    return "لا توجد بيانات"


def _get_selected_filters(request):
    # لا نختار برنامجًا افتراضيًا حتى لا تختفي باقي البرامج من الفلتر.
    return {
        "college": _safe_text(request.GET.get("college"), "all"),
        "program": _safe_text(request.GET.get("program"), "all"),
        "year": _safe_text(request.GET.get("year"), "all"),
        "report_type": _safe_text(request.GET.get("report_type"), "all"),
    }


def _get_catalog_program_metadata():
    """
    يقرأ بيانات البرامج والكليات من تطبيق programs إن كان موجودًا.
    لا تُستخدم هنا أي قائمة برامج ثابتة؛ مصدر القائمة هو قاعدة البيانات فقط.
    """
    metadata = {}

    try:
        from programs.models import Program as CatalogProgram

        catalog_programs = CatalogProgram.objects.select_related("department__college").all()

        for program in catalog_programs:
            college_name = ""
            if program.department_id and program.department and program.department.college_id:
                college_name = program.department.college.name

            display_name = _safe_text(program.name)
            if not display_name:
                continue

            item = {
                "college": college_name or DEFAULT_COLLEGE_NAME,
                "name": display_name,
            }
            metadata[_normalize_arabic(display_name)] = item
    except Exception:
        metadata = {}

    return metadata


def _program_display_name(program):
    if not program:
        return "غير محدد"

    name = _safe_text(getattr(program, "name", ""), "غير محدد")
    specialization = _safe_text(getattr(program, "specialization", ""))

    if specialization:
        return f"{name} - {specialization}"
    return name


def _program_college_name(program, catalog_metadata=None):
    if not program:
        return DEFAULT_COLLEGE_NAME

    catalog_metadata = catalog_metadata or {}
    candidates = [
        _normalize_arabic(getattr(program, "name", "")),
        _normalize_arabic(_program_display_name(program)),
    ]

    for key in candidates:
        if key in catalog_metadata:
            return catalog_metadata[key].get("college") or DEFAULT_COLLEGE_NAME

    return DEFAULT_COLLEGE_NAME


def _program_matches(program, selected_program):
    if _is_all_value(selected_program):
        return True

    selected = _normalize_arabic(selected_program)
    name = _normalize_arabic(getattr(program, "name", ""))
    specialization = _normalize_arabic(getattr(program, "specialization", ""))
    display_name = _normalize_arabic(_program_display_name(program))
    compact_display = display_name.replace(" - ", " ")

    return (
        selected == name
        or selected == specialization
        or selected == display_name
        or selected == compact_display
        or selected in display_name
        or selected in compact_display
    )



def _canonical_program_start_year(program_name):
    """
    يعيد سنة بداية البرنامج وفق القاعدة الأكاديمية المعتمدة.
    الأولوية هنا للقاعدة المعروفة حتى لو كان سجل الكتالوج لا يحتوي سنة بداية.
    """
    selected = _normalize_arabic(program_name)
    if not selected:
        return None

    normalized_map = {
        _normalize_arabic(name): year
        for name, year in PROGRAM_START_YEARS.items()
    }

    if selected in normalized_map:
        return normalized_map[selected]

    # نبحث بالأسماء الأطول أولًا حتى لا تأخذ "تقنية معلومات إنجليزي" سنة "تقنية معلومات".
    for key in sorted(normalized_map, key=len, reverse=True):
        if key and (key in selected or selected in key):
            return normalized_map[key]

    return None

def _get_dashboard_program_start_years():
    start_years = {}

    try:
        programs = AcademicProgram.objects.filter(is_active=True)
        for program in programs:
            display_name = _program_display_name(program)
            db_start_year = program.start_year or ACADEMIC_YEAR_START

            for name in [
                display_name,
                getattr(program, "name", ""),
                f"{getattr(program, 'name', '')} {getattr(program, 'specialization', '')}",
            ]:
                key = _normalize_arabic(name)
                if key:
                    start_years[key] = _canonical_program_start_year(name) or db_start_year
    except Exception:
        pass

    for name, year in PROGRAM_START_YEARS.items():
        start_years[_normalize_arabic(name)] = year

    return start_years

def _guess_start_year(program_name, dashboard_start_years=None):
    canonical_year = _canonical_program_start_year(program_name)
    if canonical_year:
        return canonical_year

    dashboard_start_years = dashboard_start_years or _get_dashboard_program_start_years()
    selected = _normalize_arabic(program_name)

    if selected in dashboard_start_years:
        return dashboard_start_years[selected]

    for key, value in sorted(dashboard_start_years.items(), key=lambda item: len(item[0]), reverse=True):
        if selected and (selected in key or key in selected):
            return value

    return ACADEMIC_YEAR_START


def _build_program_options(catalog_metadata=None):
    catalog_metadata = catalog_metadata or _get_catalog_program_metadata()
    dashboard_start_years = _get_dashboard_program_start_years()
    options = []
    seen = set()

    def should_hide_program_option(name):
        """
        يخفي خيار "نظم معلومات" العام فقط من فلتر البرامج.
        لا يخفي المسارات الفعلية التابعة له مثل:
        - نظم معلومات - ذكاء أعمال
        - نظم معلومات - أعمال إلكترونية
        """
        normalized_name = _normalize_arabic(name)
        hidden_names = {
            _normalize_arabic("نظم معلومات"),
            _normalize_arabic("نظم المعلومات"),
        }
        return normalized_name in hidden_names

    def add_option(name, college=DEFAULT_COLLEGE_NAME, start_year=ACADEMIC_YEAR_START):
        clean_name = _safe_text(name)
        if not clean_name:
            return

        if should_hide_program_option(clean_name):
            return

        key = _normalize_arabic(clean_name)
        if key in seen:
            return

        seen.add(key)
        options.append({
            "name": clean_name,
            "value": clean_name,
            "college": _safe_text(college, DEFAULT_COLLEGE_NAME),
            "start_year": int(start_year or ACADEMIC_YEAR_START),
        })

    # المصدر الأول: كتالوج البرامج الرسمي؛ لأنه يحتوي الكلية.
    try:
        from programs.models import Program as CatalogProgram
        catalog_programs = CatalogProgram.objects.select_related("department__college").all().order_by("name")
        for program in catalog_programs:
            college = DEFAULT_COLLEGE_NAME
            if program.department_id and program.department and program.department.college_id:
                college = program.department.college.name
            add_option(program.name, college, _guess_start_year(program.name, dashboard_start_years))
    except Exception:
        pass

    # المصدر الثاني: برامج ملفات التقييم؛ حتى لا تختفي البرامج التي لديها بيانات فعلية.
    try:
        programs = AcademicProgram.objects.filter(is_active=True).order_by("start_year", "name", "specialization")
        for program in programs:
            display_name = _program_display_name(program)
            add_option(display_name, _program_college_name(program, catalog_metadata), _guess_start_year(display_name, dashboard_start_years))
    except Exception:
        pass

    return sorted(options, key=lambda item: (item["start_year"], item["name"]))


def _build_college_options(program_options, selected_college=None):
    colleges = []
    seen = set()

    for option in program_options:
        college = _safe_text(option.get("college"), DEFAULT_COLLEGE_NAME)
        key = _normalize_arabic(college)
        if key and key not in seen:
            seen.add(key)
            colleges.append(college)

    if selected_college and not _is_all_value(selected_college):
        key = _normalize_arabic(selected_college)
        if key not in seen:
            colleges.append(selected_college)

    if not colleges:
        colleges.append(DEFAULT_COLLEGE_NAME)

    return colleges


def _selected_program_start_year(selected_program, program_options=None):
    """
    يحدد سنة بداية البرنامج المختار من القاعدة الأكاديمية المعتمدة.
    عند عدم اختيار برنامج محدد، يرجع أقدم سنة في النظام.
    """
    if _is_all_value(selected_program):
        return ACADEMIC_YEAR_START

    selected_key = _normalize_arabic(selected_program)

    for option in program_options or []:
        option_value = _normalize_arabic(option.get("value"))
        option_name = _normalize_arabic(option.get("name"))

        if selected_key in {option_value, option_name}:
            return int(option.get("start_year") or _guess_start_year(selected_program))

    return _guess_start_year(selected_program)


def _build_academic_year_options(selected_program="all"):
    """
    إرجاع السنوات الموجودة فعليًا في ملفات البرنامج المحدد.

    تعتمد المطابقة أولًا على الاسم الكامل للبرنامج والتخصص،
    ثم تستخدم _program_matches كحل احتياطي.
    """

    files = list(
        EvaluationFile.objects
        .select_related("program")
        .exclude(status="template_preview")
        .exclude(academic_year__isnull=True)
        .exclude(academic_year__exact="")
        .order_by("-academic_year", "-updated_at")
    )

    all_years = set()
    years_by_program = {}

    for evaluation_file in files:
        year = _safe_text(
            evaluation_file.academic_year
        ).replace("/", "-")

        if not year or year in {
            "غير محددة",
            "غير محدد",
            "-",
        }:
            continue

        program_display = _program_display_name(
            evaluation_file.program
        )

        program_key = _normalize_arabic(
            program_display
        )

        all_years.add(year)

        years_by_program.setdefault(
            program_key,
            set(),
        ).add(year)

    if _is_all_value(selected_program):
        selected_years = set(all_years)

    else:
        selected_key = _normalize_arabic(
            selected_program
        )

        selected_years = set(
            years_by_program.get(
                selected_key,
                set(),
            )
        )

        # حل احتياطي عند اختلاف كتابة الاسم أو التخصص
        if not selected_years:
            for evaluation_file in files:
                if not _program_matches(
                    evaluation_file.program,
                    selected_program,
                ):
                    continue

                year = _safe_text(
                    evaluation_file.academic_year
                ).replace("/", "-")

                if year and year not in {
                    "غير محددة",
                    "غير محدد",
                    "-",
                }:
                    selected_years.add(year)

    def year_sort_key(value):
        try:
            return int(
                _safe_text(value)
                .replace("/", "-")
                .split("-")[0]
            )
        except (TypeError, ValueError):
            return 0

    sorted_years = sorted(
        selected_years,
        key=year_sort_key,
        reverse=True,
    )

    return [
        {
            "value": year,
            "display": _display_year(year),
        }
        for year in sorted_years
    ]



def _default_year_from_options(academic_year_options):
    if not academic_year_options:
        return "all"
    return academic_year_options[0].get("value") or "all"


def _max_year_start_from_options(academic_year_options):
    starts = []
    for item in academic_year_options:
        value = _safe_text(item.get("value"))
        try:
            starts.append(int(value.split("-")[0]))
        except Exception:
            continue
    if not starts:
        return ACADEMIC_YEAR_END - 1
    return max(starts)


def _get_filtered_evaluation_files(request):
    filters = _get_selected_filters(request)
    catalog_metadata = _get_catalog_program_metadata()

    files_qs = (
        EvaluationFile.objects
        .select_related("program")
        .order_by("-updated_at")
    )

    selected_year = filters["year"]
    if not _is_all_value(selected_year):
        files_qs = files_qs.filter(academic_year=selected_year)

    files = list(files_qs)

    selected_program = filters["program"]
    if not _is_all_value(selected_program):
        files = [item for item in files if _program_matches(item.program, selected_program)]

    selected_college = filters["college"]
    if not _is_all_value(selected_college):
        selected_college_key = _normalize_arabic(selected_college)
        files = [
            item for item in files
            if _normalize_arabic(_program_college_name(item.program, catalog_metadata)) == selected_college_key
        ]

    return files, filters, catalog_metadata


def _get_program_ranking_files(request, catalog_metadata=None):
    """
    يعيد ملفات التقييم المستخدمة في ترتيب البرامج.

    يطبق فلتر السنة والكلية، لكنه يتجاهل فلتر البرنامج عمدًا؛
    حتى يحسب الترتيب العام أولًا، ثم يعرض البرنامج المختار
    برتبته الحقيقية بدل أن يصبح رقم 1 لمجرد أنه البرنامج الوحيد بعد الفلترة.
    """
    filters = _get_selected_filters(request)
    catalog_metadata = catalog_metadata or _get_catalog_program_metadata()

    files_qs = (
        EvaluationFile.objects
        .select_related("program")
        .order_by("-updated_at")
    )

    selected_year = filters["year"]
    if not _is_all_value(selected_year):
        files_qs = files_qs.filter(academic_year=selected_year)

    files = list(files_qs)

    selected_college = filters["college"]
    if not _is_all_value(selected_college):
        selected_college_key = _normalize_arabic(selected_college)
        files = [
            item
            for item in files
            if _normalize_arabic(
                _program_college_name(item.program, catalog_metadata)
            ) == selected_college_key
        ]

    return files


def _build_report_filter_summary(files, filters):
    selected_program = filters.get("program", "all")
    selected_college = filters.get("college", "all")
    selected_year = filters.get("year", "all")

    return {
        "college": "جميع الكليات" if _is_all_value(selected_college) else selected_college,
        "program": "اختار البرنامج" if _is_all_value(selected_program) else selected_program,
        "year": "جميع السنوات" if _is_all_value(selected_year) else _display_year(selected_year),
        "files_count": len(files),
    }


def _get_latest_update(files, program_reviews_qs, standard_reviews_qs, improvement_plans_qs):
    candidates = []

    for item in files:
        if item.updated_at:
            candidates.append(item.updated_at)

    for value in [
        program_reviews_qs.aggregate(value=Max("updated_at")).get("value"),
        standard_reviews_qs.aggregate(value=Max("updated_at")).get("value"),
        improvement_plans_qs.aggregate(value=Max("updated_at")).get("value"),
    ]:
        if value:
            candidates.append(value)

    if not candidates:
        return None

    return max(candidates)


def _build_standards_reports(files, file_ids):
    standards = list(QualityStandard.objects.filter(is_active=True).order_by("number"))

    entries = list(
        StandardEntry.objects
        .filter(evaluation_file_id__in=file_ids)
        .select_related("standard")
        .order_by("standard__number")
    ) if file_ids else []

    reviews = list(
        StandardEvaluationReview.objects
        .filter(review__evaluation_file_id__in=file_ids)
        .select_related("standard")
        .order_by("standard__number")
    ) if file_ids else []

    entries_by_standard = {}
    for entry in entries:
        entries_by_standard.setdefault(entry.standard_id, []).append(entry)

    reviews_by_standard = {}
    for review in reviews:
        reviews_by_standard.setdefault(review.standard_id, []).append(review)

    standards_reports = []

    for standard in standards:
        standard_entries = entries_by_standard.get(standard.id, [])
        standard_reviews = reviews_by_standard.get(standard.id, [])

        completion = _avg([entry.completion_percentage for entry in standard_entries], 0)

        score_values = []
        for review in standard_reviews:
            if review.reviewer_score is not None:
                score_values.append(review.reviewer_score)
            elif review.auto_score is not None:
                score_values.append(review.auto_score)

        score = _avg(score_values, 0)

        if not standard_entries and standard_reviews:
            percentage_values = []
            for review in standard_reviews:
                if review.reviewer_percentage is not None:
                    percentage_values.append(review.reviewer_percentage)
                else:
                    percentage_values.append(review.auto_percentage)
            completion = _avg(percentage_values, 0)

        standards_reports.append({
            "name": standard.title,
            "weight": f"{_to_int_percent(standard.weight)}%",
            "score": _format_score(score),
            "completion": _to_int_percent(completion),
            "status": _standard_status(completion, score),
        })

    return standards_reports

def _get_required_attachment_titles_by_standard():
    """
    يعيد المرفقات المطلوبة لكل معيار من نفس تعريفات صفحة إدخال البيانات.
    المرفق المتعدد يحسب كبند مطلوب واحد، مهما كان عدد الملفات داخله.
    """
    try:
        # استيراد داخل الدالة لتجنب أي circular import أثناء تشغيل Django.
        from .views import get_data_entry_standards

        standards = get_data_entry_standards()
    except Exception:
        return {}

    required_titles = {}

    for index, standard in enumerate(standards, start=1):
        try:
            standard_number = int(standard.get("number") or index)
        except (TypeError, ValueError):
            standard_number = index

        titles = []
        seen_titles = set()

        for attachment in standard.get("attachments", []):
            title = _safe_text(
                attachment.get("label") or attachment.get("name")
            )

            normalized_title = _normalize_arabic(title)

            if not normalized_title or normalized_title in seen_titles:
                continue

            seen_titles.add(normalized_title)
            titles.append(title)

        required_titles[standard_number] = titles

    return required_titles


def _build_attachment_stats(file_ids):
    """
    يحسب نسبة المرفقات حسب البنود المطلوبة، وليس حسب عدد الملفات الخام.

    مثال:
    - بند مطلوب واحد يسمح برفع عدة ملفات.
    - رفع 4 ملفات داخل نفس البند يحسب بندًا مكتملًا واحدًا فقط.
    """
    file_ids = list(dict.fromkeys(file_ids or []))
    required_titles_by_standard = (
        _get_required_attachment_titles_by_standard()
    )

    required_slots = set()

    # كل ملف تقييم يحتاج نسخة مستقلة من المرفقات المطلوبة.
    for evaluation_file_id in file_ids:
        for standard_number, titles in required_titles_by_standard.items():
            for title in titles:
                required_slots.add(
                    (
                        evaluation_file_id,
                        standard_number,
                        _normalize_arabic(title),
                    )
                )

    uploaded_slots = set()

    if file_ids:
        attachment_rows = (
            EvidenceAttachment.objects
            .filter(
                standard_entry__evaluation_file_id__in=file_ids
            )
            .values_list(
                "standard_entry__evaluation_file_id",
                "standard_entry__standard__number",
                "title",
            )
        )

        for evaluation_file_id, standard_number, title in attachment_rows:
            slot = (
                evaluation_file_id,
                standard_number,
                _normalize_arabic(title),
            )

            # لا نحسب الملفات الإضافية أو البنود غير المعرفة ضمن المطلوب.
            if slot in required_slots:
                uploaded_slots.add(slot)

    required_count = len(required_slots)
    uploaded_count = len(uploaded_slots)
    missing_count = max(required_count - uploaded_count, 0)

    rate = (
        _to_int_percent(
            (uploaded_count / required_count) * 100
        )
        if required_count
        else 0
    )

    by_standard = {}

    for standard_number, titles in required_titles_by_standard.items():
        standard_required_slots = {
            slot
            for slot in required_slots
            if slot[1] == standard_number
        }

        standard_uploaded_slots = {
            slot
            for slot in uploaded_slots
            if slot[1] == standard_number
        }

        standard_required = len(standard_required_slots)
        standard_uploaded = len(standard_uploaded_slots)

        by_standard[standard_number] = {
            "required": standard_required,
            "uploaded": standard_uploaded,
            "missing": max(
                standard_required - standard_uploaded,
                0,
            ),
        }

    return {
        "required": required_count,
        "uploaded": uploaded_count,
        "missing": missing_count,
        "rate": rate,
        "by_standard": by_standard,
    }
def _build_appendices_reports(file_ids, attachment_stats=None):
    standards = list(
        QualityStandard.objects
        .filter(is_active=True)
        .order_by("number")
    )

    attachment_stats = (
        attachment_stats
        or _build_attachment_stats(file_ids)
    )

    appendices_reports = []

    for standard in standards:
        standard_stats = attachment_stats.get(
            "by_standard",
            {},
        ).get(
            standard.number,
            {
                "required": 0,
                "uploaded": 0,
                "missing": 0,
            },
        )

        required = standard_stats["required"]
        uploaded = standard_stats["uploaded"]
        missing = standard_stats["missing"]

        if required == 0:
            status = "غير مطلوب"
        elif uploaded == 0:
            status = "غير مرفوع"
        elif missing > 0:
            status = "ناقص"
        else:
            status = "مكتمل"

        appendices_reports.append({
            "range": f"معيار {standard.number}",
            "section": standard.title,
            "required": required,
            "uploaded": uploaded,
            "missing": missing,
            "status": status,
        })

    return appendices_reports


def _build_strengths_weaknesses_reports(file_ids):
    rows = []

    if not file_ids:
        return rows

    reviews = (
        StandardEvaluationReview.objects
        .filter(review__evaluation_file_id__in=file_ids)
        .select_related("standard")
        .order_by("standard__number")
    )

    for review in reviews:
        strengths = _safe_text(review.strengths, "-")
        weaknesses = _safe_text(review.weaknesses, "-")

        if strengths == "-" and weaknesses == "-":
            continue

        rows.append({
            "standard": review.standard.title if review.standard_id else "غير محدد",
            "strengths": strengths,
            "weaknesses": weaknesses,
            "status": review.get_review_status_display() if hasattr(review, "get_review_status_display") else review.review_status,
        })

    return rows


def _build_improvement_reports(file_ids):
    rows = []

    if not file_ids:
        return rows

    plans = ImprovementPlan.objects.filter(evaluation_file_id__in=file_ids).order_by("-updated_at")

    for plan in plans:
        rows.append({
            "title": plan.title,
            "standard": plan.standard_title or (f"معيار {plan.standard_number}" if plan.standard_number else "غير محدد"),
            "action": plan.improvement_action,
            "priority": plan.get_priority_display() if hasattr(plan, "get_priority_display") else plan.priority,
            "status": plan.get_status_display() if hasattr(plan, "get_status_display") else plan.status,
            "responsible_party": plan.responsible_party or "-",
            "due_date": plan.due_date.strftime("%Y/%m/%d") if plan.due_date else "-",
        })

    return rows


def _build_program_rankings(files, selected_program="all"):
    """
    يرتب البرامج حسب نسبة الجاهزية الشاملة نفسها المستخدمة في
    "التقرير الشامل للبرنامج":

    - اكتمال بيانات المعايير: 45%
    - نتائج التقييم: 35%
    - المرفقات والشواهد: 20%

    تبقى ready وmissing لعرض عدد المعايير المكتملة وغير المكتملة،
    لكنها لا تكون أساس الترتيب. كما يحسب الترتيب العام قبل تطبيق
    فلتر البرنامج حتى لا يتحول كل برنامج مختار منفردًا إلى المرتبة الأولى.
    """
    files_by_program = {}

    for evaluation_file in files:
        if not evaluation_file.program_id or not evaluation_file.program:
            continue

        files_by_program.setdefault(
            evaluation_file.program_id,
            [],
        ).append(evaluation_file)

    rankings = []

    for program_files in files_by_program.values():
        program = program_files[0].program
        file_ids = [item.id for item in program_files]

        entries_qs = StandardEntry.objects.filter(
            evaluation_file_id__in=file_ids
        )

        required = entries_qs.count()
        ready = entries_qs.filter(
            completion_percentage__gte=100
        ).count()
        missing = max(required - ready, 0)

        completion_rate = _to_int_percent(
            _avg(
                entries_qs.values_list(
                    "completion_percentage",
                    flat=True,
                ),
                0,
            )
        )

        program_reviews_qs = ProgramEvaluationReview.objects.filter(
            evaluation_file_id__in=file_ids
        )

        final_percentages = []
        for review in program_reviews_qs:
            if review.final_percentage is not None:
                final_percentages.append(review.final_percentage)
            elif review.overall_reviewer_percentage is not None:
                final_percentages.append(
                    review.overall_reviewer_percentage
                )
            else:
                final_percentages.append(
                    review.overall_auto_percentage
                )

        evaluation_progress = _to_int_percent(
            _avg(final_percentages, 0)
        )

        attachment_stats = _build_attachment_stats(file_ids)
        attachments_rate = attachment_stats["rate"]

        comprehensive_progress = _to_int_percent(
            (completion_rate * 0.45)
            + (evaluation_progress * 0.35)
            + (attachments_rate * 0.20)
        )

        rankings.append({
            "program": _program_display_name(program),
            "program_id": program.id,
            "ready": ready,
            "missing": missing,
            "required": required,
            "progress": comprehensive_progress,
            "missing_progress": max(
                0,
                100 - comprehensive_progress,
            ),
            "missing_percent": max(
                0,
                100 - comprehensive_progress,
            ),
            "completion_rate": completion_rate,
            "evaluation_progress": evaluation_progress,
            "attachments_rate": attachments_rate,
            "_program_object": program,
        })

    rankings.sort(
        key=lambda item: (
            -item["progress"],
            -item["ready"],
            _normalize_arabic(item["program"]),
        )
    )

    for index, item in enumerate(rankings, start=1):
        item["rank"] = index

    if not _is_all_value(selected_program):
        rankings = [
            item
            for item in rankings
            if _program_matches(
                item["_program_object"],
                selected_program,
            )
        ]
    else:
        rankings = rankings[:8]

    for item in rankings:
        item.pop("_program_object", None)

    return rankings


def _build_report_types(report_progress):
    """
    يبني حالة كل تقرير من نسبة جاهزيته.

    كان الخطأ السابق يعيّن جميع التقارير غير تقرير المرفقات إلى
    "لا توجد بيانات" مهما كانت نسبة الجاهزية، بسبب موضع else.
    """
    report_types = []

    for definition in REPORT_DEFINITIONS:
        report_id = definition["id"]
        progress = _to_int_percent(
            report_progress.get(report_id, 0)
        )

        if report_id == "attachments":
            if progress >= 100:
                status, status_class = "مكتمل", "success"
            elif progress > 0:
                status, status_class = "مسودة", "warning"
            else:
                status, status_class = "لا توجد بيانات", "danger"
        else:
            status, status_class = _status_from_progress(progress)

        report_types.append({
            "id": report_id,
            "title": definition["title"],
            "description": definition["description"],
            "status": status,
            "status_class": status_class,
            "progress": progress,
        })

    return report_types


def _build_status_counts(report_types):
    ready = len([item for item in report_types if item.get("status_class") == "success"])
    in_progress = len([item for item in report_types if item.get("status_class") == "info"])
    draft = len([item for item in report_types if item.get("status_class") == "warning"])
    no_data = len([item for item in report_types if item.get("status_class") == "danger"])

    return {
        "ready": ready,
        "review": in_progress,
        "in_progress": in_progress,
        "draft": draft,
        "empty": no_data,
        "no_data": no_data,
    }



def _build_recent_export_logs(limit=20):
    """
    يعرض آخر التقارير التي تم تصديرها فعليًا.
    إذا لم تكن migration مطبقة بعد، يرجع قائمة فارغة بدل تعطيل صفحة التقارير.
    """
    try:
        logs = (
            ReportExportLog.objects
            .select_related("exported_by")
            .order_by("-created_at")[:limit]
        )
    except Exception:
        return []

    recent_logs = []
    for log in logs:
        report_type = _safe_text(log.report_type, "all")
        college = _safe_text(log.college_name, "جميع الكليات")
        program = _safe_text(log.program_name, "اختار البرنامج")
        year = _safe_text(log.academic_year, "جميع السنوات")

        query = {
            "college": "all" if college in {"جميع الكليات", ""} else college,
            "program": "all" if program in {"اختار البرنامج", "جميع البرامج", ""} else program,
            "year": "all" if year in {"جميع السنوات", ""} else year.replace("/", "-"),
            "report_type": report_type if report_type in REPORT_TITLES or report_type == "all" else "all",
        }

        if log.export_format == "PDF":
            format_label = "PDF"
            format_class = "file-pdf"
        elif log.export_format == "EXCEL":
            format_label = "XLS"
            format_class = "file-xls"
        else:
            format_label = _safe_text(log.export_format, "DOC")
            format_class = "file-doc"

        exported_by = ""
        if getattr(log, "exported_by_id", None) and log.exported_by:
            exported_by = _safe_text(log.exported_by.get_username())

        recent_logs.append({
            "id": log.id,
            "report_id": report_type,
            "title": _safe_text(log.report_title, _selected_report_title(report_type)),
            "program": program,
            "college": college,
            "year": _display_year(year),
            "created_at": _format_date(log.created_at),
            "format": log.export_format,
            "format_label": format_label,
            "format_class": format_class,
            "exported_by": exported_by,
            "query_string": urlencode(query),
        })

    return recent_logs


def _log_report_export(request, data, export_format):
    """
    يسجل عملية التصدير في جدول ReportExportLog فقط إذا كان التقرير يحتوي على بيانات فعلية.
    لا يتم تسجيل التقارير الفارغة في قسم التقارير الحديثة.
    """
    try:
        selected_type = _selected_report_type(request)

        if not _has_meaningful_report_data(data, selected_type):
            return False

        filter_summary = data.get("filter_summary", {}) if isinstance(data, dict) else {}
        user = getattr(request, "user", None)
        if not getattr(user, "is_authenticated", False):
            user = None

        ReportExportLog.objects.create(
            report_type=selected_type,
            report_title=_selected_report_title(selected_type),
            college_name=_safe_text(filter_summary.get("college"), "جميع الكليات"),
            program_name=_safe_text(filter_summary.get("program"), "اختار البرنامج"),
            academic_year=_safe_text(filter_summary.get("year"), "جميع السنوات"),
            export_format=export_format,
            exported_by=user,
        )
        return True
    except Exception:
        # لا نريد أن يفشل تحميل التقرير بسبب مشكلة في جدول السجل أو migration غير مطبقة.
        return False

def _build_reports_data_from_database(request):
    files, filters, catalog_metadata = _get_filtered_evaluation_files(request)
    file_ids = [item.id for item in files]

    program_reviews_qs = ProgramEvaluationReview.objects.filter(evaluation_file_id__in=file_ids) if file_ids else ProgramEvaluationReview.objects.none()
    standard_reviews_qs = StandardEvaluationReview.objects.filter(review__evaluation_file_id__in=file_ids) if file_ids else StandardEvaluationReview.objects.none()
    improvement_plans_qs = ImprovementPlan.objects.filter(evaluation_file_id__in=file_ids) if file_ids else ImprovementPlan.objects.none()
    standard_entries_qs = StandardEntry.objects.filter(evaluation_file_id__in=file_ids) if file_ids else StandardEntry.objects.none()

    completion_rate = _to_int_percent(_avg(standard_entries_qs.values_list("completion_percentage", flat=True), 0))

    attachment_stats = _build_attachment_stats(file_ids)

    required_attachments = attachment_stats["required"]
    uploaded_attachments = attachment_stats["uploaded"]
    attachments_rate = attachment_stats["rate"]

    final_percentages = []
    for review in program_reviews_qs:
        if review.final_percentage is not None:
            final_percentages.append(review.final_percentage)
        elif review.overall_reviewer_percentage is not None:
            final_percentages.append(review.overall_reviewer_percentage)
        else:
            final_percentages.append(review.overall_auto_percentage)

    evaluation_progress = _to_int_percent(_avg(final_percentages, 0))

    reviewed_standards_count = standard_reviews_qs.filter(review_status="reviewed").count()
    total_standards_reviews_count = standard_reviews_qs.count()
    strengths_weaknesses_progress = _to_int_percent((reviewed_standards_count / total_standards_reviews_count) * 100) if total_standards_reviews_count else 0

    plans_total = improvement_plans_qs.count()
    plans_closed = improvement_plans_qs.filter(status__in=["completed", "closed"]).count()
    plans_open = improvement_plans_qs.filter(status__in=["proposed", "in_progress"]).count()
    improvement_progress = _to_int_percent((plans_closed / plans_total) * 100) if plans_total else 0

    comprehensive_progress = _to_int_percent((completion_rate * 0.45) + (evaluation_progress * 0.35) + (attachments_rate * 0.20))
    executive_progress = _to_int_percent((comprehensive_progress * 0.70) + (improvement_progress * 0.30))

    report_progress = {
        "comprehensive": comprehensive_progress,
        "evaluation": evaluation_progress,
        "attachments": attachments_rate,
        "strengths_weaknesses": strengths_weaknesses_progress,
        "improvement": improvement_progress,
        "executive": executive_progress,
    }

    report_types = _build_report_types(report_progress)
    ready_reports = len([report for report in report_types if report["status_class"] == "success"])
    status_counts = _build_status_counts(report_types)

    standards_reports = _build_standards_reports(files, file_ids)
    appendices_reports = _build_appendices_reports(
        file_ids,
        attachment_stats,
    )
    strengths_weaknesses_reports = _build_strengths_weaknesses_reports(file_ids)
    improvement_reports = _build_improvement_reports(file_ids)

    program_options = _build_program_options(catalog_metadata)
    college_options = _build_college_options(program_options, filters.get("college"))
    selected_program_start_year = _selected_program_start_year(filters.get("program"), program_options)
    academic_year_options = _build_academic_year_options(
        filters.get("program")
    )
    academic_years = [item["value"] for item in academic_year_options]
    default_academic_year = _default_year_from_options(academic_year_options)
    max_academic_year_start = _max_year_start_from_options(academic_year_options)
    ranking_files = _get_program_ranking_files(
        request,
        catalog_metadata,
    )
    program_rankings = _build_program_rankings(
        ranking_files,
        filters.get("program", "all"),
    )

    latest_update = _get_latest_update(files, program_reviews_qs, standard_reviews_qs, improvement_plans_qs)

    return {
        "report_summary": {
            "completion_rate": completion_rate,
            "ready_reports": ready_reports,
            "attachments_rate": attachments_rate,
            "open_improvement_plans": plans_open,
            "files_count": len(files),
            "uploaded_attachments": uploaded_attachments,
            "required_attachments": required_attachments,
            "latest_update": _format_date(latest_update),
        },
        "report_types": report_types,
        "standards_reports": standards_reports,
        "appendices_reports": appendices_reports,
        "strengths_weaknesses_reports": strengths_weaknesses_reports,
        "improvement_reports": improvement_reports,
        "program_rankings": program_rankings,
        "programs": program_options,
        "colleges": college_options,
        "academic_years": academic_years,
        "academic_year_options": academic_year_options,
        "selected_filters": filters,
        "filter_summary": _build_report_filter_summary(files, filters),
        "status_counts": status_counts,
        "default_academic_year": default_academic_year,
        "max_academic_year_start": max_academic_year_start,
        "has_real_data": bool(files),
        "reports_data_error": "",
        "recent_exports": _build_recent_export_logs(),
    }


def _empty_reports_context(request, message="", technical_error=None):
    filters = _get_selected_filters(request)

    report_types = []
    for definition in REPORT_DEFINITIONS:
        report_types.append({
            "id": definition["id"],
            "title": definition["title"],
            "description": definition["description"],
            "status": "لا توجد بيانات",
            "status_class": "danger",
            "progress": 0,
        })

    try:
        catalog_metadata = _get_catalog_program_metadata()
        program_options = _build_program_options(catalog_metadata)
        college_options = _build_college_options(program_options, filters.get("college"))
        selected_program_start_year = _selected_program_start_year(filters.get("program"), program_options)
        academic_year_options = _build_academic_year_options(filters.get("program"))
    except Exception:
        program_options = []
        college_options = []
        academic_year_options = []

    default_academic_year = _default_year_from_options(academic_year_options)
    max_academic_year_start = _max_year_start_from_options(academic_year_options)

    return {
        "report_summary": {
            "completion_rate": 0,
            "ready_reports": 0,
            "attachments_rate": 0,
            "open_improvement_plans": 0,
            "files_count": 0,
            "uploaded_attachments": 0,
            "required_attachments": 0,
            "latest_update": "لا توجد بيانات",
        },
        "report_types": report_types,
        "standards_reports": [],
        "appendices_reports": [],
        "strengths_weaknesses_reports": [],
        "improvement_reports": [],
        "program_rankings": [],
        "programs": program_options,
        "colleges": college_options,
        "academic_years": [item["value"] for item in academic_year_options],
        "academic_year_options": academic_year_options,
        "selected_filters": filters,
        "filter_summary": {
            "college": "جميع الكليات" if _is_all_value(filters.get("college")) else filters.get("college"),
            "program": "اختار البرنامج" if _is_all_value(filters.get("program")) else filters.get("program"),
            "year": "جميع السنوات" if _is_all_value(filters.get("year")) else _display_year(filters.get("year")),
            "files_count": 0,
        },
        "status_counts": _build_status_counts(report_types),
        "default_academic_year": default_academic_year,
        "max_academic_year_start": max_academic_year_start,
        "has_real_data": False,
        "reports_data_error": message,
        "reports_data_technical_error": _safe_text(technical_error),
        "recent_exports": _build_recent_export_logs(),
    }


def get_reports_data(request=None):
    """
    مصدر بيانات واجهة التقارير من قاعدة البيانات مع الحفاظ على مفاتيح القالب القديمة.
    """
    request = request or type("Request", (), {"GET": {}})()

    try:
        return _build_reports_data_from_database(request)
    except DatabaseError as error:
        return _empty_reports_context(
            request,
            "تعذر قراءة بيانات التقارير من قاعدة البيانات. تأكدي من تنفيذ migrations ثم إعادة تشغيل السيرفر.",
            technical_error=error,
        )
    except Exception as error:
        return _empty_reports_context(
            request,
            "حدث خطأ أثناء تجهيز بيانات التقارير.",
            technical_error=error,
        )


def reports(request):
    context = get_reports_data(request)
    context["current_full_path"] = request.get_full_path()
    return render(request, "dashboard/reports.html", context)


@require_POST
def delete_report_export_log(request, log_id):
    """
    يحذف سجلًا واحدًا من قسم التقارير الحديثة فقط.
    لا يحذف ملفات PDF/Excel ولا يؤثر على بيانات التقارير الأصلية.
    """
    try:
        ReportExportLog.objects.filter(id=log_id).delete()
    except Exception:
        pass

    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER")

    if not next_url or not url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        next_url = reverse("reports")

    return redirect(next_url)


def _selected_report_type(request):
    report_type = _safe_text(request.GET.get("report_type"), "all")
    if report_type not in REPORT_TITLES and report_type != "all":
        return "all"
    return report_type


def _selected_report_title(report_type):
    if report_type == "all":
        return "جميع التقارير"
    return REPORT_TITLES.get(report_type, "جميع التقارير")


def _has_meaningful_report_data(data, selected_type):
    """
    يمنع تصدير التقارير الفارغة ومنع تسجيلها في سجل التقارير الحديثة.
    وجود ملف تقييم فقط لا يعني أن كل أنواع التقارير لديها بيانات.
    لذلك نتحقق من نوع التقرير المطلوب تحديدًا.
    """
    if not isinstance(data, dict):
        return False

    summary = data.get("report_summary", {}) or {}
    has_files = bool(data.get("has_real_data")) or int(summary.get("files_count") or 0) > 0

    standards = data.get("standards_reports") or []
    appendices = data.get("appendices_reports") or []
    strengths = data.get("strengths_weaknesses_reports") or []
    improvements = data.get("improvement_reports") or []

    def has_evaluation_data():
        for row in standards:
            status = _safe_text(row.get("status"))
            score = _safe_text(row.get("score"))
            completion = row.get("completion") or 0
            try:
                completion = float(completion)
            except (TypeError, ValueError):
                completion = 0
            if status != "لا توجد بيانات" and (completion > 0 or score not in {"", "-", "0", "0.0"}):
                return True
        return False

    def has_attachments_data():
        for row in appendices:
            required = int(row.get("required") or 0)
            uploaded = int(row.get("uploaded") or 0)
            missing = int(row.get("missing") or 0)
            status = _safe_text(row.get("status"))
            if status != "لا توجد بيانات" and (required > 0 or uploaded > 0 or missing > 0):
                return True
        return False

    def has_strengths_data():
        for row in strengths:
            strength = _safe_text(row.get("strengths"))
            weakness = _safe_text(row.get("weaknesses"))
            if strength not in {"", "-", "لا توجد بيانات"} or weakness not in {"", "-", "لا توجد بيانات"}:
                return True
        return False

    def has_improvement_data():
        return bool(improvements)

    checks = {
        "evaluation": has_evaluation_data,
        "attachments": has_attachments_data,
        "strengths_weaknesses": has_strengths_data,
        "improvement": has_improvement_data,
    }

    if selected_type in checks:
        return checks[selected_type]()

    if selected_type == "executive":
        return has_evaluation_data() or has_improvement_data()

    if selected_type in {"all", "comprehensive"}:
        return has_files and (
            has_evaluation_data()
            or has_attachments_data()
            or has_strengths_data()
            or has_improvement_data()
            or int(summary.get("uploaded_attachments") or 0) > 0
            or int(summary.get("required_attachments") or 0) > 0
        )

    return False


def _export_no_data_response(selected_type):
    report_title = _selected_report_title(selected_type)
    html = f"""
    <!doctype html>
    <html lang="ar" dir="rtl">
    <head>
        <meta charset="utf-8">
        <title>لا توجد بيانات للتصدير</title>
        <style>
            body {{ font-family: Tahoma, Arial, sans-serif; background:#f8fafc; color:#0b2a4a; margin:0; }}
            .box {{ max-width: 720px; margin: 80px auto; background:#fff; border:1px solid #dbe5f0; border-radius:18px; padding:28px; line-height:1.9; box-shadow:0 12px 30px rgba(15,23,42,.08); }}
            h1 {{ margin:0 0 12px; font-size:22px; }}
            p {{ margin:0 0 10px; color:#475569; }}
            .note {{ background:#fff7ed; border:1px solid #fed7aa; color:#9a3412; padding:12px 14px; border-radius:12px; margin-top:16px; }}
            button {{ margin-top:18px; padding:10px 18px; border:0; border-radius:10px; background:#0b63ce; color:#fff; cursor:pointer; font-weight:700; }}
        </style>
    </head>
    <body>
        <div class="box">
            <h1>لا توجد بيانات قابلة للتصدير</h1>
            <p>التقرير المحدد: <strong>{report_title}</strong></p>
            <div class="note">يمكن تغيير البرنامج أو الكلية أو السنة الأكاديمية، أو اختيار نوع تقرير آخر يحتوي على بيانات.</div>
            <button onclick="history.back()">رجوع</button>
        </div>
    </body>
    </html>
    """
    return HttpResponse(html, status=400)


def _excel_style_helpers(workbook):
    main_color = "1F4568"
    light_fill = "EDF4FB"
    white_color = "FFFFFF"
    border_color = "D9E2EC"

    header_fill = PatternFill("solid", fgColor=main_color)
    section_fill = PatternFill("solid", fgColor=light_fill)

    header_font = Font(color=white_color, bold=True, size=12)
    title_font = Font(color=main_color, bold=True, size=16)
    section_font = Font(color=main_color, bold=True, size=13)
    normal_font = Font(size=11)

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    def style_all_cells(sheet):
        sheet.sheet_view.rightToLeft = True

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = right_alignment
                cell.font = normal_font
                cell.border = thin_border

        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(value))
            sheet.column_dimensions[column_letter].width = min(max_length + 5, 50)

    def style_header_row(sheet, row_number):
        for cell in sheet[row_number]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

    def style_section_row(sheet, row_number):
        for cell in sheet[row_number]:
            cell.fill = section_fill
            cell.font = section_font
            cell.alignment = right_alignment

    return title_font, center_alignment, style_all_cells, style_header_row, style_section_row


def _add_excel_title(sheet, title, subtitle, title_font, center_alignment, width="F"):
    sheet.merge_cells(f"A1:{width}1")
    sheet["A1"] = title
    sheet["A1"].font = title_font
    sheet["A1"].alignment = center_alignment
    sheet.merge_cells(f"A2:{width}2")
    sheet["A2"] = subtitle
    sheet["A2"].alignment = center_alignment
    sheet.append([""])


def _add_summary_sheet(workbook, data, title_font, center_alignment, style_all_cells, style_header_row):
    summary = data["report_summary"]
    filter_summary = data.get("filter_summary", {})
    sheet = workbook.create_sheet("ملخص التقرير")
    _add_excel_title(sheet, "ملخص التقارير", f"البرنامج: {filter_summary.get('program')} | السنة: {filter_summary.get('year')}", title_font, center_alignment, "B")
    sheet.append(["البند", "القيمة"])
    style_header_row(sheet, sheet.max_row)
    sheet.append(["الكلية", filter_summary.get("college", "جميع الكليات")])
    sheet.append(["البرنامج", filter_summary.get("program", "اختار البرنامج")])
    sheet.append(["السنة الأكاديمية", filter_summary.get("year", "جميع السنوات")])
    sheet.append(["ملفات البيانات المطابقة", summary.get("files_count", 0)])
    sheet.append(["مؤشر جاهزية التقرير العام", f"{summary['completion_rate']}%"])
    sheet.append(["عدد التقارير الجاهزة", summary["ready_reports"]])
    sheet.append(["نسبة اكتمال المرفقات", f"{summary['attachments_rate']}%"])
    sheet.append(["عدد خطط التحسين المفتوحة", summary["open_improvement_plans"]])
    sheet.append(["آخر تحديث", summary.get("latest_update", "لا توجد بيانات")])
    style_all_cells(sheet)
    sheet.freeze_panes = "A4"


def _add_report_types_sheet(workbook, data, title_font, center_alignment, style_all_cells, style_header_row):
    sheet = workbook.create_sheet("أنواع التقارير")
    _add_excel_title(sheet, "أنواع التقارير", "حالة كل نوع تقرير حسب الفلاتر الحالية", title_font, center_alignment, "E")
    sheet.append(["م", "نوع التقرير", "الوصف", "الحالة", "نسبة الجاهزية"])
    style_header_row(sheet, sheet.max_row)
    for index, report in enumerate(data["report_types"], start=1):
        sheet.append([index, report["title"], report["description"], report["status"], f"{report['progress']}%"])
    style_all_cells(sheet)
    sheet.freeze_panes = "A4"


def _add_standards_sheet(workbook, data, title_font, center_alignment, style_all_cells, style_header_row):
    sheet = workbook.create_sheet("نتائج التقييم")
    _add_excel_title(sheet, "تقرير نتائج التقييم", "حالة المعايير ونسب الاكتمال والدرجات", title_font, center_alignment, "F")
    sheet.append(["م", "المعيار", "الوزن النسبي", "درجة التقييم من 5", "نسبة الاكتمال", "الحالة"])
    style_header_row(sheet, sheet.max_row)
    for index, standard in enumerate(data["standards_reports"], start=1):
        sheet.append([index, standard["name"], standard["weight"], standard["score"], f"{standard['completion']}%", standard["status"]])
    style_all_cells(sheet)
    sheet.freeze_panes = "A4"


def _add_appendices_sheet(workbook, data, title_font, center_alignment, style_all_cells, style_header_row):
    sheet = workbook.create_sheet("المرفقات والشواهد")
    _add_excel_title(sheet, "تقرير المرفقات والشواهد", "حالة الشواهد المرتبطة بكل معيار", title_font, center_alignment, "G")
    sheet.append(["م", "نطاق الملاحق", "القسم المرتبط", "المطلوب", "المرفوع", "الناقص", "الحالة"])
    style_header_row(sheet, sheet.max_row)
    for index, appendix in enumerate(data["appendices_reports"], start=1):
        sheet.append([index, appendix["range"], appendix["section"], appendix["required"], appendix["uploaded"], appendix["missing"], appendix["status"]])
    style_all_cells(sheet)
    sheet.freeze_panes = "A4"


def _add_strengths_sheet(workbook, data, title_font, center_alignment, style_all_cells, style_header_row):
    sheet = workbook.create_sheet("القوة والضعف")
    _add_excel_title(sheet, "تقرير نقاط القوة والضعف", "النقاط المسجلة في مراجعة المعايير", title_font, center_alignment, "E")
    sheet.append(["م", "المعيار", "نقاط القوة", "نقاط الضعف", "الحالة"])
    style_header_row(sheet, sheet.max_row)
    rows = data.get("strengths_weaknesses_reports", [])
    if rows:
        for index, row in enumerate(rows, start=1):
            sheet.append([index, row["standard"], row["strengths"], row["weaknesses"], row["status"]])
    else:
        sheet.append([1, "-", "لا توجد بيانات مسجلة", "لا توجد بيانات مسجلة", "لا توجد بيانات"])
    style_all_cells(sheet)
    sheet.freeze_panes = "A4"


def _add_improvement_sheet(workbook, data, title_font, center_alignment, style_all_cells, style_header_row):
    sheet = workbook.create_sheet("خطط التحسين")
    _add_excel_title(sheet, "تقرير خطط التحسين", "خطط التحسين المرتبطة بالفلاتر الحالية", title_font, center_alignment, "G")
    sheet.append(["م", "الخطة", "المعيار", "الإجراء التحسيني", "الأولوية", "الحالة", "المسؤول"])
    style_header_row(sheet, sheet.max_row)
    rows = data.get("improvement_reports", [])
    if rows:
        for index, row in enumerate(rows, start=1):
            sheet.append([index, row["title"], row["standard"], row["action"], row["priority"], row["status"], row["responsible_party"]])
    else:
        sheet.append([1, "لا توجد خطط تحسين", "-", "-", "-", "لا توجد بيانات", "-"])
    style_all_cells(sheet)
    sheet.freeze_panes = "A4"


def export_reports_excel(request):
    data = get_reports_data(request)
    selected_type = _selected_report_type(request)

    if not _has_meaningful_report_data(data, selected_type):
        return _export_no_data_response(selected_type)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    title_font, center_alignment, style_all_cells, style_header_row, style_section_row = _excel_style_helpers(workbook)

    if selected_type in ["all", "comprehensive", "executive"]:
        _add_summary_sheet(workbook, data, title_font, center_alignment, style_all_cells, style_header_row)

    if selected_type in ["all", "comprehensive"]:
        _add_report_types_sheet(workbook, data, title_font, center_alignment, style_all_cells, style_header_row)

    if selected_type in ["all", "comprehensive", "evaluation", "executive"]:
        _add_standards_sheet(workbook, data, title_font, center_alignment, style_all_cells, style_header_row)

    if selected_type in ["all", "comprehensive", "attachments"]:
        _add_appendices_sheet(workbook, data, title_font, center_alignment, style_all_cells, style_header_row)

    if selected_type in ["all", "comprehensive", "strengths_weaknesses"]:
        _add_strengths_sheet(workbook, data, title_font, center_alignment, style_all_cells, style_header_row)

    if selected_type in ["all", "comprehensive", "improvement", "executive"]:
        _add_improvement_sheet(workbook, data, title_font, center_alignment, style_all_cells, style_header_row)

    workbook.active = 0

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"academic_reports_{selected_type}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    _log_report_export(request, data, "EXCEL")
    return response


def _register_arabic_font():
    font_paths = [
        r"C:\Windows\Fonts\tahoma.ttf",
        r"C:\Windows\Fonts\Tahoma.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ]

    font_path = None
    for path in font_paths:
        if os.path.exists(path):
            font_path = path
            break

    if not font_path:
        return None

    if "TahomaArabic" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("TahomaArabic", font_path))

    return "TahomaArabic"


def export_reports_pdf(request):
    data = get_reports_data(request)
    selected_type = _selected_report_type(request)
    report_type_name = _selected_report_title(selected_type)
    filter_summary = data.get("filter_summary", {})

    if not _has_meaningful_report_data(data, selected_type):
        return _export_no_data_response(selected_type)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="academic_reports_{selected_type}.pdf"'

    font_name = _register_arabic_font()
    if not font_name:
        return HttpResponse(
            "لم يتم العثور على خط عربي مناسب لإنشاء ملف PDF. تأكدي من وجود Tahoma على ويندوز أو DejaVuSans على الاستضافة.",
            status=500,
        )

    def ar(value):
        value = "" if value is None else str(value)
        reshaped_text = arabic_reshaper.reshape(value)
        return get_display(reshaped_text)

    title_style = ParagraphStyle(
        name="ArabicTitle",
        fontName=font_name,
        fontSize=18,
        leading=24,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#1F4568"),
        spaceAfter=12,
    )

    subtitle_style = ParagraphStyle(
        name="ArabicSubtitle",
        fontName=font_name,
        fontSize=11,
        leading=16,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#475569"),
        spaceAfter=10,
    )

    section_style = ParagraphStyle(
        name="ArabicSection",
        fontName=font_name,
        fontSize=13,
        leading=18,
        alignment=TA_RIGHT,
        textColor=colors.white,
    )

    normal_style = ParagraphStyle(
        name="ArabicNormal",
        fontName=font_name,
        fontSize=9,
        leading=13,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
    )

    right_style = ParagraphStyle(
        name="ArabicRight",
        fontName=font_name,
        fontSize=9,
        leading=13,
        alignment=TA_RIGHT,
        textColor=colors.HexColor("#111827"),
    )

    def p(value, style=None):
        return Paragraph(ar(value), style or normal_style)

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.1 * cm,
        bottomMargin=1.1 * cm,
        title=ar(report_type_name),
    )

    story = []
    story.append(Paragraph(ar(report_type_name), title_style))
    story.append(Paragraph(ar("نظام الجودة والأداء الأكاديمي"), subtitle_style))
    story.append(Spacer(1, 10))

    info_table = Table(
        [
            [
                p(f"الكلية: {filter_summary.get('college', 'جميع الكليات')}", right_style),
                p(f"البرنامج: {filter_summary.get('program', 'اختار البرنامج')}", right_style),
            ],
            [
                p(f"السنة الأكاديمية: {filter_summary.get('year', 'جميع السنوات')}", right_style),
                p(f"نوع التقرير: {report_type_name}", right_style),
            ],
        ],
        colWidths=[13 * cm, 13 * cm],
    )

    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#D9E2EC")),
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))

    story.append(info_table)
    story.append(Spacer(1, 12))

    def add_section_title(title):
        section_table = Table([[Paragraph(ar(title), section_style)]], colWidths=[26 * cm])
        section_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1F4568")),
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#1F4568")),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(section_table)
        story.append(Spacer(1, 8))

    def add_table(rows, widths):
        table = Table(rows, colWidths=widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EDF4FB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1F4568")),
            ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#D9E2EC")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))

    def add_summary_section():
        summary = data["report_summary"]
        add_section_title("ملخص التقرير")
        add_table(
            [
                [p("خطط التحسين المفتوحة"), p("نسبة اكتمال المرفقات"), p("التقارير الجاهزة"), p("ملفات البيانات")],
                [p(summary["open_improvement_plans"]), p(f"{summary['attachments_rate']}%"), p(summary["ready_reports"]), p(summary.get("files_count", 0))],
            ],
            [6.5 * cm, 6.5 * cm, 6.5 * cm, 6.5 * cm],
        )

    def add_report_types_section():
        add_section_title("أنواع التقارير")
        rows = [[p("نسبة الجاهزية"), p("الحالة"), p("الوصف"), p("نوع التقرير"), p("م")]]
        for index, report in enumerate(data["report_types"], start=1):
            rows.append([p(f"{report['progress']}%"), p(report["status"]), p(report["description"]), p(report["title"]), p(index)])
        add_table(rows, [3 * cm, 3 * cm, 9 * cm, 8 * cm, 3 * cm])

    def add_standards_section():
        add_section_title("نتائج التقييم")
        rows = [[p("الحالة"), p("نسبة الاكتمال"), p("درجة التقييم من 5"), p("الوزن النسبي"), p("المعيار"), p("م")]]
        for index, standard in enumerate(data["standards_reports"], start=1):
            rows.append([p(standard["status"]), p(f"{standard['completion']}%"), p(standard["score"]), p(standard["weight"]), p(standard["name"]), p(index)])
        if len(rows) == 1:
            rows.append([p("لا توجد بيانات"), p("-"), p("-"), p("-"), p("-"), p(1)])
        add_table(rows, [4 * cm, 3.5 * cm, 3.8 * cm, 3.2 * cm, 9 * cm, 2.5 * cm])

    def add_appendices_section():
        add_section_title("المرفقات والشواهد")
        rows = [[p("الحالة"), p("الناقص"), p("المرفوع"), p("المطلوب"), p("القسم المرتبط"), p("نطاق الملاحق"), p("م")]]
        for index, appendix in enumerate(data["appendices_reports"], start=1):
            rows.append([p(appendix["status"]), p(appendix["missing"]), p(appendix["uploaded"]), p(appendix["required"]), p(appendix["section"]), p(appendix["range"]), p(index)])
        if len(rows) == 1:
            rows.append([p("لا توجد بيانات"), p("-"), p("-"), p("-"), p("-"), p("-"), p(1)])
        add_table(rows, [3.5 * cm, 2.7 * cm, 2.7 * cm, 2.7 * cm, 6 * cm, 6 * cm, 2.4 * cm])

    def add_strengths_section():
        add_section_title("نقاط القوة والضعف")
        rows = [[p("الحالة"), p("نقاط الضعف"), p("نقاط القوة"), p("المعيار"), p("م")]]
        items = data.get("strengths_weaknesses_reports", [])
        if items:
            for index, row in enumerate(items, start=1):
                rows.append([p(row["status"]), p(row["weaknesses"], right_style), p(row["strengths"], right_style), p(row["standard"]), p(index)])
        else:
            rows.append([p("لا توجد بيانات"), p("-"), p("-"), p("-"), p(1)])
        add_table(rows, [3 * cm, 8 * cm, 8 * cm, 5 * cm, 2 * cm])

    def add_improvement_section():
        add_section_title("خطط التحسين")
        rows = [[p("الحالة"), p("الأولوية"), p("المسؤول"), p("الإجراء التحسيني"), p("الخطة"), p("م")]]
        items = data.get("improvement_reports", [])
        if items:
            for index, row in enumerate(items, start=1):
                rows.append([p(row["status"]), p(row["priority"]), p(row["responsible_party"]), p(row["action"], right_style), p(row["title"]), p(index)])
        else:
            rows.append([p("لا توجد بيانات"), p("-"), p("-"), p("-"), p("لا توجد خطط تحسين"), p(1)])
        add_table(rows, [3 * cm, 3 * cm, 4 * cm, 9 * cm, 5 * cm, 2 * cm])

    if selected_type in ["all", "comprehensive", "executive"]:
        add_summary_section()

    if selected_type in ["all", "comprehensive"]:
        add_report_types_section()

    if selected_type in ["all", "comprehensive", "evaluation", "executive"]:
        add_standards_section()

    if selected_type in ["all", "comprehensive", "attachments"]:
        add_appendices_section()

    if selected_type in ["all", "comprehensive", "strengths_weaknesses"]:
        add_strengths_section()

    if selected_type in ["all", "comprehensive", "improvement", "executive"]:
        add_improvement_section()

    note_table = Table([[p("تم إنشاء هذا الملف تلقائيًا حسب الفلاتر ونوع التقرير المحدد في واجهة التقارير.", right_style)]], colWidths=[26 * cm])
    note_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFFDF4")),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#F1DF9A")),
        ("PADDING", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
    ]))
    story.append(note_table)

    try:
        doc.build(story)
    except Exception as error:
        return HttpResponse(f"حدث خطأ أثناء إنشاء ملف PDF: {error}", status=500)

    _log_report_export(request, data, "PDF")
    return response
